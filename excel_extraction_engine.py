"""Enterprise-grade, config-driven Excel extraction engine.

This module is intentionally free of any Databricks runtime dependency
(no ``dbutils``, ``spark``, or ``display``). It can therefore be unit-tested
in isolation, packaged into a wheel, and reused by multiple notebooks or jobs.

Design goals addressed (vs. the original notebooks):

* No hardcoded domain tokens. The "index" key column(s) and the side-by-side
  "section" dimension name are configurable instead of being literally
  ``"Year"`` / ``"Sector"``.
* Lossless, configurable type handling. Text-stored numbers are coerced only
  when it is safe (whole-column, leading-zero codes preserved, locale-aware
  thousands/decimal separators), with per-column overrides.
* Merged cells are resolved so banner/section/header cells read correctly.
* Multi-row (hierarchical) headers are supported.
* Header-name collisions are de-duplicated rather than silently overwritten.
* End-of-table is detected by blank-row / footer markers rather than scanning
  to the last non-null cell (which used to swallow footnotes).
* Configurable scan windows replace the magic ``10`` / ``100`` / ``500`` limits.
* Strict config validation with actionable error messages.
* Structured logging, per-table metrics, run id and config hash for lineage.
* Atomic writes and output-name de-duplication.

The public surface is:

* :class:`ExtractionSettings`        – workbook-level behaviour
* :func:`validate_config`            – fail-fast schema validation
* :class:`WorkbookReader`            – merged-cell-aware openpyxl wrapper
* :class:`ExcelExtractor`            – the extraction orchestrator
* :class:`SheetAnalyser`             – structure detection for config generation
* :func:`build_config`              – assemble a config from analyses
* :class:`ExtractionWriter`          – parquet/csv writer + manifest
* :func:`normalize_local_path`       – dbfs/Volumes path handling
"""

from __future__ import annotations

import hashlib
import json
import logging
import os
import re
import uuid
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import openpyxl
import pandas as pd
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.utils.exceptions import InvalidFileException

__all__ = [
    "ENGINE_VERSION",
    "ExtractionSettings",
    "ExtractionError",
    "ConfigValidationError",
    "UnsupportedFileError",
    "validate_config",
    "WorkbookReader",
    "ValueCleaner",
    "HeaderResolver",
    "ExcelExtractor",
    "TableResult",
    "SheetAnalyser",
    "build_config",
    "ExtractionWriter",
    "normalize_local_path",
    "utc_now_iso",
]

ENGINE_VERSION = "2.0.0"

logger = logging.getLogger("excel_extraction")
if not logger.handlers:
    # Library default; the host notebook/app can reconfigure freely.
    _h = logging.StreamHandler()
    _h.setFormatter(logging.Formatter("%(asctime)s | %(levelname)-7s | %(message)s"))
    logger.addHandler(_h)
    logger.setLevel(logging.INFO)


# --------------------------------------------------------------------------- #
# Exceptions
# --------------------------------------------------------------------------- #
class ExtractionError(Exception):
    """Base class for engine errors."""


class ConfigValidationError(ExtractionError):
    """Raised when a config fails validation."""


class UnsupportedFileError(ExtractionError):
    """Raised for file formats openpyxl cannot read (e.g. legacy ``.xls``)."""


# --------------------------------------------------------------------------- #
# Helpers
# --------------------------------------------------------------------------- #
def utc_now_iso() -> str:
    """Timezone-aware UTC timestamp (``datetime.utcnow`` is deprecated)."""
    return datetime.now(timezone.utc).isoformat()


def normalize_local_path(path: str) -> str:
    """Translate Databricks storage URIs to local-file-API paths where possible.

    * ``dbfs:/x``    -> ``/dbfs/x`` (DBFS FUSE mount)
    * ``/Volumes/x`` -> unchanged (UC Volumes FUSE)
    * ``abfss://``/``s3://``/``wasbs://`` are returned unchanged and will only
      work if the runtime has a FUSE mount; otherwise the caller should stage
      the file locally first.
    """
    if not path:
        return path
    if path.startswith("dbfs:/"):
        return "/dbfs/" + path[len("dbfs:/"):].lstrip("/")
    return path


def stable_config_hash(config: Dict[str, Any]) -> str:
    """Deterministic short hash of a config for lineage/audit."""
    blob = json.dumps(config, sort_keys=True, default=str).encode("utf-8")
    return hashlib.sha256(blob).hexdigest()[:16]


def _col_to_idx(letter_or_int: Any) -> int:
    """Accept a column letter (``"A"``) or 1-based int and return a 1-based int."""
    if isinstance(letter_or_int, int):
        return letter_or_int
    return column_index_from_string(str(letter_or_int).strip())


# --------------------------------------------------------------------------- #
# Settings
# --------------------------------------------------------------------------- #
# NB: a bare "-" is intentionally NOT a default null marker. In the original
# code it nulled legitimate dashes. Statistical-disclosure markers remain.
DEFAULT_NULL_MARKERS: Tuple[str, ...] = ("[x]", "[c]", "[z]", "..", "")
DEFAULT_NOTE_PATTERN = r"\[[Nn]ote\s*\d+\]"


@dataclass
class ExtractionSettings:
    """Workbook-level behaviour, overridable via the config ``settings`` block."""

    null_markers: frozenset = field(default_factory=lambda: frozenset(DEFAULT_NULL_MARKERS))
    note_pattern: str = DEFAULT_NOTE_PATTERN
    type_inference: str = "safe"          # "none" | "safe" | "aggressive"
    decimal_sep: str = "."
    thousands_sep: str = ","
    header_search_max_rows: int = 25
    max_scan_cols: int = 1000
    max_scan_rows: int = 1_048_576        # Excel hard limit
    stop_on_blank_rows: int = 1           # end a table after N consecutive blank rows
    footer_markers: Tuple[str, ...] = ()  # e.g. ("Source:", "Notes")
    resolve_merged_cells: bool = True
    dedupe_headers: bool = True
    on_error: str = "continue"            # "continue" | "fail_fast"

    @classmethod
    def from_config(cls, config: Dict[str, Any]) -> "ExtractionSettings":
        """Build settings, supporting both v2 (``settings`` block) and v1
        (top-level ``null_markers`` / ``note_pattern``) configs."""
        s = dict(config.get("settings", {}))
        # v1 backward-compat: promote top-level keys if a settings block is absent.
        for legacy in ("null_markers", "note_pattern"):
            if legacy in config and legacy not in s:
                s[legacy] = config[legacy]

        null_markers = s.get("null_markers")
        kwargs: Dict[str, Any] = {}
        if null_markers is not None:
            kwargs["null_markers"] = frozenset(null_markers)
        for k in (
            "note_pattern", "type_inference", "decimal_sep", "thousands_sep",
            "header_search_max_rows", "max_scan_cols", "max_scan_rows",
            "stop_on_blank_rows", "resolve_merged_cells", "dedupe_headers",
            "on_error",
        ):
            if k in s:
                kwargs[k] = s[k]
        if "footer_markers" in s:
            kwargs["footer_markers"] = tuple(s["footer_markers"])
        inst = cls(**kwargs)
        if inst.type_inference not in ("none", "safe", "aggressive"):
            raise ConfigValidationError(
                f"settings.type_inference must be one of none|safe|aggressive, "
                f"got {inst.type_inference!r}")
        if inst.on_error not in ("continue", "fail_fast"):
            raise ConfigValidationError(
                f"settings.on_error must be continue|fail_fast, got {inst.on_error!r}")
        return inst


# --------------------------------------------------------------------------- #
# Config validation
# --------------------------------------------------------------------------- #
_VALID_TABLE_TYPES = {"simple", "side_by_side", "stacked"}


def _require(cond: bool, msg: str) -> None:
    if not cond:
        raise ConfigValidationError(msg)


def validate_config(config: Dict[str, Any]) -> None:
    """Validate a config, raising :class:`ConfigValidationError` on first problem.

    This is deliberately strict so that misconfiguration fails fast at the top
    of a run rather than producing silently-wrong tables.
    """
    _require(isinstance(config, dict), "Config must be a JSON object.")
    _require("config_name" in config and config["config_name"],
             "Config must include a non-empty 'config_name'.")
    _require(isinstance(config.get("sheets"), list) and config["sheets"],
             "Config must include a non-empty 'sheets' list.")
    # Validate settings early (raises on bad enums).
    ExtractionSettings.from_config(config)

    seen_output_names: Dict[str, str] = {}
    for i, sc in enumerate(config["sheets"]):
        loc = f"sheets[{i}]"
        _require(isinstance(sc, dict), f"{loc} must be an object.")
        _require(sc.get("sheet_name"), f"{loc} missing 'sheet_name'.")
        _require(sc.get("output_name"), f"{loc} ({sc.get('sheet_name')}) missing 'output_name'.")
        tt = sc.get("table_type")
        _require(tt in _VALID_TABLE_TYPES,
                 f"{loc} ({sc['sheet_name']}) has invalid table_type {tt!r}; "
                 f"expected one of {sorted(_VALID_TABLE_TYPES)}.")

        out = sc["output_name"]
        if out in seen_output_names:
            raise ConfigValidationError(
                f"Duplicate output_name {out!r} on sheets "
                f"{seen_output_names[out]!r} and {sc['sheet_name']!r}; "
                f"output names must be unique.")
        seen_output_names[out] = sc["sheet_name"]

        # Header rows: accept single header_row or list header_rows.
        if tt in ("simple", "side_by_side"):
            has_single = "header_row" in sc
            has_multi = "header_rows" in sc
            _require(has_single or has_multi,
                     f"{loc} ({sc['sheet_name']}) needs 'header_row' or 'header_rows'.")
            if has_multi:
                _require(isinstance(sc["header_rows"], list) and sc["header_rows"],
                         f"{loc} 'header_rows' must be a non-empty list of ints.")
            _require("data_start_row" in sc,
                     f"{loc} ({sc['sheet_name']}) needs 'data_start_row'.")

        if tt == "side_by_side":
            _require(isinstance(sc.get("sections"), list) and sc["sections"],
                     f"{loc} ({sc['sheet_name']}) side_by_side needs a non-empty 'sections' list.")
            for j, sec in enumerate(sc["sections"]):
                _require(sec.get("name"), f"{loc}.sections[{j}] missing 'name'.")
                _require(sec.get("start_col") and sec.get("end_col"),
                         f"{loc}.sections[{j}] needs 'start_col' and 'end_col'.")
                try:
                    if _col_to_idx(sec["start_col"]) > _col_to_idx(sec["end_col"]):
                        raise ConfigValidationError(
                            f"{loc}.sections[{j}] start_col is right of end_col.")
                except KeyError as e:
                    raise ConfigValidationError(f"{loc}.sections[{j}] invalid column: {e}")

        if tt == "stacked":
            _require(isinstance(sc.get("tables"), list) and sc["tables"],
                     f"{loc} ({sc['sheet_name']}) stacked needs a non-empty 'tables' list.")
            for j, tbl in enumerate(sc["tables"]):
                _require("header_row" in tbl and "data_start_row" in tbl,
                         f"{loc}.tables[{j}] needs 'header_row' and 'data_start_row'.")

        if "columns" in sc:
            for j, cd in enumerate(sc["columns"]):
                _require(cd.get("col") and cd.get("name"),
                         f"{loc}.columns[{j}] needs 'col' and 'name'.")


# --------------------------------------------------------------------------- #
# Workbook reader (merged-cell aware)
# --------------------------------------------------------------------------- #
class WorkbookReader:
    """openpyxl wrapper that resolves merged cells and reports formula gaps."""

    def __init__(self, file_path: str, resolve_merged_cells: bool = True):
        self.file_path = normalize_local_path(file_path)
        self.resolve_merged_cells = resolve_merged_cells
        ext = os.path.splitext(self.file_path)[1].lower()
        if ext in (".xls", ".xlsb"):
            raise UnsupportedFileError(
                f"File format {ext!r} is not supported by this engine. "
                f"openpyxl reads .xlsx/.xlsm only. Convert the file to .xlsx "
                f"(e.g. via Excel 'Save As', or a one-off pandas/libreoffice "
                f"conversion) and re-run.")
        try:
            # data_only=True returns the last-cached computed value of formulas.
            self.wb = openpyxl.load_workbook(self.file_path, data_only=True, read_only=False)
        except InvalidFileException as e:
            raise UnsupportedFileError(f"Could not open {self.file_path!r}: {e}")
        self._merged_cache: Dict[str, Dict[Tuple[int, int], Any]] = {}

    @property
    def sheet_names(self) -> List[str]:
        return list(self.wb.sheetnames)

    def has_sheet(self, name: str) -> bool:
        return name in self.wb.sheetnames

    def _merged_map(self, ws) -> Dict[Tuple[int, int], Any]:
        """Map every cell inside a merged range to its top-left value."""
        if not self.resolve_merged_cells:
            return {}
        if ws.title in self._merged_cache:
            return self._merged_cache[ws.title]
        m: Dict[Tuple[int, int], Any] = {}
        for rng in ws.merged_cells.ranges:
            top_left = ws.cell(row=rng.min_row, column=rng.min_col).value
            if top_left is None:
                continue
            for r in range(rng.min_row, rng.max_row + 1):
                for c in range(rng.min_col, rng.max_col + 1):
                    if (r, c) != (rng.min_row, rng.min_col):
                        m[(r, c)] = top_left
        self._merged_cache[ws.title] = m
        return m

    def cell(self, ws, row: int, col: int) -> Any:
        """Return a cell value, resolving merged ranges."""
        v = ws.cell(row=row, column=col).value
        if v is None and self.resolve_merged_cells:
            return self._merged_map(ws).get((row, col))
        return v

    def worksheet(self, name: str):
        return self.wb[name]

    def dims(self, ws) -> Tuple[int, int]:
        return ws.max_row, ws.max_column

    def close(self) -> None:
        try:
            self.wb.close()
        except Exception:  # pragma: no cover - defensive
            pass

    def __enter__(self) -> "WorkbookReader":
        return self

    def __exit__(self, *exc) -> None:
        self.close()


# --------------------------------------------------------------------------- #
# Value cleaning & type coercion
# --------------------------------------------------------------------------- #
class ValueCleaner:
    """Null handling, string normalisation and locale-aware type coercion."""

    _LEADING_ZERO = re.compile(r"^0\d+$")

    def __init__(self, settings: ExtractionSettings):
        self.s = settings
        self._note_re = re.compile(settings.note_pattern) if settings.note_pattern else None

    # -- headers --
    def clean_header(self, val: Any) -> Optional[str]:
        if val is None:
            return None
        s = str(val).strip()
        if self._note_re:
            s = self._note_re.sub("", s)
        s = re.sub(r"\s+", " ", s.replace("\n", " ")).strip()
        return s or None

    # -- values (null + strip only; typing happens column-wise later) --
    def clean_value(self, val: Any) -> Any:
        if val is None:
            return None
        if isinstance(val, str):
            s = val.strip()
            if s in self.s.null_markers:
                return None
            return s if s != "" else None
        return val

    # -- column-wise numeric coercion --
    def _try_number(self, s: str) -> Optional[Any]:
        """Parse a single string to int/float using configured separators.

        Returns ``None`` if not safely numeric. Leading-zero codes (``"007"``)
        are treated as non-numeric to preserve identifiers.
        """
        t = s.strip()
        if t == "":
            return None
        if self._LEADING_ZERO.match(t):
            return None  # identifier/code, keep as string
        cleaned = t
        if self.s.thousands_sep:
            cleaned = cleaned.replace(self.s.thousands_sep, "")
        if self.s.decimal_sep and self.s.decimal_sep != ".":
            cleaned = cleaned.replace(self.s.decimal_sep, ".")
        # Reject anything with stray non-numeric chars (currency, %, units).
        if not re.fullmatch(r"[+-]?\d*\.?\d+(?:[eE][+-]?\d+)?", cleaned):
            return None
        try:
            if "." in cleaned or "e" in cleaned or "E" in cleaned:
                return float(cleaned)
            return int(cleaned)
        except ValueError:
            return None

    def coerce_column(self, series: pd.Series, override: Optional[str]) -> pd.Series:
        """Coerce a column's dtype according to ``type_inference`` and override.

        ``override`` is one of ``string|int|float|number|keep`` or ``None``.
        """
        if override == "keep":
            return series
        if override == "string":
            out = series.astype("object").where(series.notna(), None)
            return out.map(lambda v: None if v is None else str(v))

        non_null = series.dropna()
        if non_null.empty:
            return series

        # Identify string cells that *could* be numbers.
        str_mask = non_null.map(lambda v: isinstance(v, str))

        if override in ("int", "float", "number"):
            mode = "aggressive"
        else:
            mode = self.s.type_inference
        if mode == "none":
            return series

        if mode == "safe":
            # Convert only if EVERY string cell parses as a number.
            str_vals = non_null[str_mask]
            if str_vals.empty:
                return series  # already-typed numbers/dates; leave as-is
            parsed = str_vals.map(self._try_number)
            if parsed.isna().any():
                return series  # mixed -> keep strings to avoid silent loss
            converted = series.map(
                lambda v: self._try_number(v) if isinstance(v, str) else v)
            return self._finalize_numeric(converted, override)

        # aggressive: parse each string individually, keep non-parseable as str
        converted = series.map(
            lambda v: (self._try_number(v) if isinstance(v, str)
                       and self._try_number(v) is not None else v))
        return self._finalize_numeric(converted, override)

    @staticmethod
    def _finalize_numeric(series: pd.Series, override: Optional[str]) -> pd.Series:
        non_null = series.dropna()
        if non_null.empty:
            return series
        all_numeric = non_null.map(lambda v: isinstance(v, (int, float))).all()
        if override == "float":
            return series.map(lambda v: float(v) if isinstance(v, (int, float)) else v)
        if all_numeric:
            all_int = non_null.map(
                lambda v: isinstance(v, int) or (isinstance(v, float) and float(v).is_integer())
            ).all()
            if all_int and override != "float":
                # Use pandas nullable integer to keep NULLs without float coercion.
                try:
                    return series.astype("Int64")
                except (TypeError, ValueError):
                    return series
        return series


# --------------------------------------------------------------------------- #
# Header resolution
# --------------------------------------------------------------------------- #
class HeaderResolver:
    """Builds clean, de-duplicated, optionally multi-row headers."""

    def __init__(self, reader: WorkbookReader, cleaner: ValueCleaner,
                 settings: ExtractionSettings):
        self.reader = reader
        self.cleaner = cleaner
        self.s = settings

    def header_rows_for(self, sheet_config: Dict[str, Any]) -> List[int]:
        if "header_rows" in sheet_config:
            return [int(r) for r in sheet_config["header_rows"]]
        return [int(sheet_config["header_row"])]

    def build(self, ws, start_col: int, end_col: int,
              header_rows: Sequence[int], join: str = " | ") -> List[Optional[str]]:
        """Return one header label per column in ``[start_col, end_col]``."""
        labels: List[Optional[str]] = []
        for c in range(start_col, end_col + 1):
            parts: List[str] = []
            for hr in header_rows:
                val = self.cleaner.clean_header(self.reader.cell(ws, hr, c))
                if val and (not parts or parts[-1] != val):
                    parts.append(val)
            labels.append(join.join(parts) if parts else None)
        if self.s.dedupe_headers:
            labels = self._dedupe(labels)
        return labels

    @staticmethod
    def _dedupe(labels: List[Optional[str]]) -> List[Optional[str]]:
        seen: Dict[str, int] = {}
        out: List[Optional[str]] = []
        for lab in labels:
            if lab is None:
                out.append(None)
                continue
            if lab in seen:
                seen[lab] += 1
                out.append(f"{lab}_{seen[lab]}")
            else:
                seen[lab] = 1
                out.append(lab)
        return out


# --------------------------------------------------------------------------- #
# Extraction result
# --------------------------------------------------------------------------- #
@dataclass
class TableResult:
    output_name: str
    sheet_name: str
    table_type: str
    dataframe: pd.DataFrame
    status: str = "ok"            # "ok" | "empty" | "error"
    error: Optional[str] = None
    duration_ms: float = 0.0

    @property
    def rows(self) -> int:
        return 0 if self.dataframe is None else len(self.dataframe)

    @property
    def cols(self) -> int:
        return 0 if self.dataframe is None else len(self.dataframe.columns)


# --------------------------------------------------------------------------- #
# Extractor
# --------------------------------------------------------------------------- #
class ExcelExtractor:
    """Config-driven extractor. Returns :class:`TableResult` objects.

    ``unpivot_mode`` overrides per-sheet config: ``as_configured`` (default),
    ``force_wide``, or ``force_unpivot``.
    """

    def __init__(self, config: Dict[str, Any], unpivot_mode: str = "as_configured"):
        validate_config(config)
        self.config = config
        self.settings = ExtractionSettings.from_config(config)
        self.cleaner = ValueCleaner(self.settings)
        if unpivot_mode not in ("as_configured", "force_wide", "force_unpivot"):
            raise ConfigValidationError(
                f"unpivot_mode must be as_configured|force_wide|force_unpivot, "
                f"got {unpivot_mode!r}")
        self.unpivot_mode = unpivot_mode
        self.run_id = str(uuid.uuid4())
        self.config_hash = stable_config_hash(config)
        self.log_records: List[Dict[str, Any]] = []

    # -- logging --
    def _log(self, sheet: str, message: str, level: str = "INFO") -> None:
        self.log_records.append({
            "timestamp": utc_now_iso(), "run_id": self.run_id,
            "sheet": sheet, "level": level, "message": message,
        })
        getattr(logger, level.lower() if level != "OK" else "info")(f"[{sheet}] {message}")

    # -- geometry helpers --
    def _last_data_col(self, reader: WorkbookReader, ws, row: int, start_col: int) -> int:
        _, max_col = reader.dims(ws)
        last = start_col
        for c in range(start_col, min(self.settings.max_scan_cols, max_col) + 1):
            if reader.cell(ws, row, c) is not None:
                last = c
        return last

    def _is_footer(self, reader: WorkbookReader, ws, row: int,
                   start_col: int, end_col: int) -> bool:
        if not self.settings.footer_markers:
            return False
        first = reader.cell(ws, row, start_col)
        if first is None:
            return False
        text = str(first).strip()
        return any(text.startswith(m) for m in self.settings.footer_markers)

    def _row_blank(self, reader: WorkbookReader, ws, row: int,
                   start_col: int, end_col: int) -> bool:
        for c in range(start_col, end_col + 1):
            if reader.cell(ws, row, c) is not None:
                return False
        return True

    def _data_end_row(self, reader: WorkbookReader, ws, start_row: int,
                      start_col: int, end_col: int,
                      explicit_end: Optional[int]) -> int:
        max_row, _ = reader.dims(ws)
        hard_end = min(explicit_end or max_row, max_row, self.settings.max_scan_rows)
        if explicit_end:
            return hard_end
        last_data = start_row - 1
        consecutive_blanks = 0
        for r in range(start_row, hard_end + 1):
            if self._is_footer(reader, ws, r, start_col, end_col):
                break
            if self._row_blank(reader, ws, r, start_col, end_col):
                consecutive_blanks += 1
                if consecutive_blanks >= self.settings.stop_on_blank_rows:
                    break
            else:
                consecutive_blanks = 0
                last_data = r
        return max(last_data, start_row - 1)

    # -- core range extraction --
    def _extract_range(self, reader: WorkbookReader, ws, headers: List[Optional[str]],
                       start_col: int, data_start: int, data_end: int,
                       index_cols: Sequence[str]) -> pd.DataFrame:
        rows: List[Dict[str, Any]] = []
        index_set = set(index_cols)
        for r in range(data_start, data_end + 1):
            row_data: Dict[str, Any] = {}
            has_payload = False
            for offset, name in enumerate(headers):
                if not name:
                    continue
                val = self.cleaner.clean_value(reader.cell(ws, r, start_col + offset))
                row_data[name] = val
                # A row counts as data if any *non-index* column is populated.
                if val is not None and name not in index_set:
                    has_payload = True
            if has_payload or (row_data and not index_set):
                rows.append(row_data)
        df = pd.DataFrame(rows)
        # Ensure all header columns exist even if entirely blank in the slice.
        for name in headers:
            if name and name not in df.columns:
                df[name] = None
        ordered = [h for h in headers if h]
        return df[ordered] if ordered else df

    # -- post-processing --
    def _index_cols(self, sheet_config: Dict[str, Any]) -> List[str]:
        return list(sheet_config.get("index_columns", []))

    def _should_unpivot(self, sheet_config: Dict[str, Any]) -> bool:
        if self.unpivot_mode == "force_wide":
            return False
        if self.unpivot_mode == "force_unpivot":
            return True
        return bool(sheet_config.get("unpivot", False))

    def _unpivot(self, df: pd.DataFrame, sheet_config: Dict[str, Any]) -> pd.DataFrame:
        if df.empty:
            return df
        var_name = sheet_config.get("unpivot_var_name", "Variable")
        val_name = sheet_config.get("unpivot_value_name", "Value")
        exclude = set(sheet_config.get("exclude_from_unpivot", []))
        section_col = sheet_config.get("section_column_name")

        if "unpivot_id_cols" in sheet_config:
            id_cols = [c for c in sheet_config["unpivot_id_cols"] if c in df.columns]
        else:
            id_cols = [c for c in self._index_cols(sheet_config) if c in df.columns]
            if section_col and section_col in df.columns and section_col not in id_cols:
                id_cols.append(section_col)
        all_id = id_cols + [c for c in df.columns if c in exclude and c not in id_cols]
        value_cols = [c for c in df.columns if c not in all_id and c is not None]
        if not value_cols:
            return df
        melted = df.melt(id_vars=all_id, value_vars=value_cols,
                         var_name=var_name, value_name=val_name)
        return melted.dropna(subset=[val_name])

    def _coerce_types(self, df: pd.DataFrame, sheet_config: Dict[str, Any]) -> pd.DataFrame:
        if df.empty:
            return df
        overrides = sheet_config.get("column_types", {})
        for col in df.columns:
            df[col] = self.cleaner.coerce_column(df[col], overrides.get(col))
        return df

    def _post_process(self, df: pd.DataFrame, sheet_config: Dict[str, Any]) -> pd.DataFrame:
        if df is None:
            return pd.DataFrame()
        # Drop configured columns (supports "Prefix*" wildcard).
        for col in sheet_config.get("drop_columns", []):
            if col in df.columns:
                df = df.drop(columns=[col])
            elif "*" in col:
                stem = col.replace("*", "")
                df = df.drop(columns=[c for c in df.columns if c and stem in str(c)],
                             errors="ignore")
        # Drop all-null spacer columns.
        df = df.dropna(axis=1, how="all")
        # Rename (newline/whitespace tolerant).
        rename_map = sheet_config.get("rename_columns", {})
        if rename_map:
            norm = {}
            for old, new in rename_map.items():
                norm[old] = new
                norm[re.sub(r"\s+", " ", str(old).replace("\n", " ")).strip()] = new
            actual = {}
            for col in df.columns:
                key = col if col in norm else re.sub(r"\s+", " ", str(col).replace("\n", " ")).strip()
                if key in norm:
                    actual[col] = norm[key]
            if actual:
                df = df.rename(columns=actual)
        # Static columns.
        for name, val in sheet_config.get("static_columns", {}).items():
            df[name] = val
        # Type coercion BEFORE unpivot so numbers melt as numbers.
        df = self._coerce_types(df, sheet_config)
        # Unpivot.
        if self._should_unpivot(sheet_config):
            df = self._unpivot(df, sheet_config)
        return df.reset_index(drop=True)

    # -- table-type handlers --
    def _resolve_cols(self, sheet_config: Dict[str, Any], reader: WorkbookReader,
                      ws, header_rows: List[int]) -> Tuple[int, int]:
        start_col = _col_to_idx(sheet_config.get("start_col", "A"))
        end_col = self._last_data_col(reader, ws, header_rows[-1], start_col)
        return start_col, end_col

    def _extract_simple(self, reader: WorkbookReader, ws,
                        sheet_config: Dict[str, Any], hr: HeaderResolver) -> pd.DataFrame:
        header_rows = hr.header_rows_for(sheet_config)
        data_start = int(sheet_config["data_start_row"])
        index_cols = self._index_cols(sheet_config)

        if "columns" in sheet_config:
            col_defs = sheet_config["columns"]
            first_idx = _col_to_idx(col_defs[0]["col"])
            data_end = self._data_end_row(reader, ws, data_start, first_idx,
                                          self._last_data_col(reader, ws, header_rows[-1], first_idx),
                                          sheet_config.get("data_end_row"))
            rows = []
            names = [cd["name"] for cd in col_defs]
            idxs = [_col_to_idx(cd["col"]) for cd in col_defs]
            for r in range(data_start, data_end + 1):
                row_data, has_payload = {}, False
                for name, ci in zip(names, idxs):
                    val = self.cleaner.clean_value(reader.cell(ws, r, ci))
                    row_data[name] = val
                    if val is not None and name not in index_cols:
                        has_payload = True
                if has_payload:
                    rows.append(row_data)
            df = pd.DataFrame(rows, columns=names)
        else:
            start_col, end_col = self._resolve_cols(sheet_config, reader, ws, header_rows)
            headers = hr.build(ws, start_col, end_col, header_rows,
                               sheet_config.get("header_join", " | "))
            data_end = self._data_end_row(reader, ws, data_start, start_col, end_col,
                                          sheet_config.get("data_end_row"))
            df = self._extract_range(reader, ws, headers, start_col, data_start,
                                     data_end, index_cols)
        return self._post_process(df, sheet_config)

    def _extract_side_by_side(self, reader: WorkbookReader, ws,
                              sheet_config: Dict[str, Any], hr: HeaderResolver) -> pd.DataFrame:
        header_rows = hr.header_rows_for(sheet_config)
        data_start = int(sheet_config["data_start_row"])
        index_cols = self._index_cols(sheet_config)
        section_col = sheet_config.get("section_column_name", "Section")
        sections = sheet_config["sections"]

        first_start = _col_to_idx(sections[0]["start_col"])
        first_end = _col_to_idx(sections[0]["end_col"])
        data_end = self._data_end_row(reader, ws, data_start, first_start, first_end,
                                      sheet_config.get("data_end_row"))
        frames = []
        for section in sections:
            sc = _col_to_idx(section["start_col"])
            ec = _col_to_idx(section["end_col"])
            headers = hr.build(ws, sc, ec, header_rows,
                               sheet_config.get("header_join", " | "))
            sub = self._extract_range(reader, ws, headers, sc, data_start, data_end, index_cols)
            if sub.empty:
                self._log(sheet_config["sheet_name"],
                          f"Section '{section['name']}' empty", "WARN")
                continue
            sub[section_col] = section["name"]
            frames.append(sub)
            self._log(sheet_config["sheet_name"],
                      f"Section '{section['name']}': {len(sub)} rows")
        combined = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
        return self._post_process(combined, sheet_config)

    def _extract_stacked(self, reader: WorkbookReader, ws,
                         sheet_config: Dict[str, Any], hr: HeaderResolver) -> pd.DataFrame:
        index_cols = self._index_cols(sheet_config)
        frames = []
        for tbl in sheet_config["tables"]:
            header_rows = [int(r) for r in tbl.get("header_rows", [tbl["header_row"]])]
            data_start = int(tbl["data_start_row"])
            start_col = _col_to_idx(tbl.get("start_col", "A"))
            end_col = self._last_data_col(reader, ws, header_rows[-1], start_col)
            data_end = self._data_end_row(reader, ws, data_start, start_col, end_col,
                                          tbl.get("data_end_row"))
            headers = hr.build(ws, start_col, end_col, header_rows,
                               sheet_config.get("header_join", " | "))
            sub = self._extract_range(reader, ws, headers, start_col, data_start,
                                      data_end, index_cols)
            for name, val in tbl.get("static_columns", {}).items():
                sub[name] = val
            frames.append(sub)
            self._log(sheet_config["sheet_name"],
                      f"Stacked table '{tbl.get('name', '?')}': {len(sub)} rows")
        combined = pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
        return self._post_process(combined, sheet_config)

    # -- orchestration --
    def extract_all(self, file_path: str,
                    sheet_filter: Optional[Sequence[str]] = None) -> Dict[str, TableResult]:
        import time
        results: Dict[str, TableResult] = {}
        reader = WorkbookReader(file_path, self.settings.resolve_merged_cells)
        hr = HeaderResolver(reader, self.cleaner, self.settings)
        available = set(reader.sheet_names)
        try:
            if sheet_filter:
                missing = [s for s in sheet_filter if s not in available]
                if missing:
                    self._log("FILTER", f"Not in workbook: {missing}", "WARN")
                    self._log("FILTER", f"Available: {reader.sheet_names}", "INFO")

            for sc in self.config["sheets"]:
                name = sc["sheet_name"]
                out = sc["output_name"]
                tt = sc["table_type"]
                if sheet_filter and name not in sheet_filter:
                    continue
                if name not in available:
                    self._log(name, "Sheet not found — skipping", "WARN")
                    continue

                t0 = time.perf_counter()
                ws = reader.worksheet(name)
                try:
                    self._log(name, f"Extracting ({tt})...")
                    if tt == "simple":
                        df = self._extract_simple(reader, ws, sc, hr)
                    elif tt == "side_by_side":
                        df = self._extract_side_by_side(reader, ws, sc, hr)
                    else:
                        df = self._extract_stacked(reader, ws, sc, hr)
                    dt = (time.perf_counter() - t0) * 1000
                    status = "empty" if df.empty else "ok"
                    results[out] = TableResult(out, name, tt, df, status, None, dt)
                    self._log(name, f"Extracted {len(df)} rows x {len(df.columns)} cols", "OK")
                except Exception as e:  # noqa: BLE001
                    dt = (time.perf_counter() - t0) * 1000
                    if self.settings.on_error == "fail_fast":
                        raise ExtractionError(f"[{name}] extraction failed: {e}") from e
                    self._log(name, f"Extraction failed: {e}", "ERROR")
                    results[out] = TableResult(out, name, tt, pd.DataFrame(),
                                               "error", str(e), dt)
        finally:
            reader.close()
        return results


# --------------------------------------------------------------------------- #
# Structure analysis (for config generation)
# --------------------------------------------------------------------------- #
class SheetAnalyser:
    """Best-effort structure detection. Output is a *starting point* for a
    config that a human should review — not an authoritative parse."""

    def __init__(self, reader: WorkbookReader, sheet_name: str,
                 settings: ExtractionSettings, index_keyword: str = "year"):
        self.reader = reader
        self.ws = reader.worksheet(sheet_name)
        self.name = sheet_name
        self.s = settings
        self.cleaner = ValueCleaner(settings)
        self.index_keyword = index_keyword.strip().lower()
        self.max_row, self.max_col = reader.dims(self.ws)

    def _filled(self, row: int) -> Dict[int, Any]:
        vals = {}
        for c in range(1, min(self.s.max_scan_cols, self.max_col) + 1):
            v = self.reader.cell(self.ws, row, c)
            if v is not None:
                vals[c] = v
        return vals

    def _count(self, row: int) -> int:
        return len(self._filled(row))

    def _last_data_row(self, col: int = 1) -> int:
        last = 1
        for r in range(1, min(self.s.max_scan_rows, self.max_row) + 1):
            if self.reader.cell(self.ws, r, col) is not None:
                last = r
        return last

    def _max_col_in_row(self, row: int) -> int:
        last = 1
        for c in range(1, min(self.s.max_scan_cols, self.max_col) + 1):
            if self.reader.cell(self.ws, row, c) is not None:
                last = c
        return last

    def _label_count(self, row: int) -> int:
        """Count genuine text-label cells: non-null strings that are neither a
        null marker nor a number.

        Header rows are overwhelmingly text labels; data rows are numbers and
        null markers. This separates them robustly even in *side-by-side*
        layouts, where the header labels repeat across sections (so a
        distinct-value count would wrongly favour a data row full of unique
        numbers) and in files with merged banner rows above the header."""
        n = 0
        for v in self._filled(row).values():
            if isinstance(v, str):
                t = v.strip()
                if t and t not in self.s.null_markers and self.cleaner._try_number(t) is None:
                    n += 1
        return n

    def detect_header_row(self) -> Tuple[Optional[int], List[Tuple[int, int]]]:
        candidates = []
        window = min(self.s.header_search_max_rows, self.max_row)
        for r in range(1, window + 1):
            if self._count(r) >= 2:
                nxt = self._count(r + 1) if r < self.max_row else 0
                if nxt >= 2:
                    candidates.append((r, self._label_count(r), self._count(r)))
        if not candidates:
            return None, []
        # Most text labels wins; tie-break by total filled cells, then the
        # later row (a column-header row sits directly above the first data row,
        # below any banner/title rows).
        best = max(candidates, key=lambda x: (x[1], x[2], x[0]))
        return best[0], [(r, lc) for r, lc, _ in candidates]

    def detect_section_header_row(self, header_row: int) -> Optional[int]:
        if header_row <= 1:
            return None
        header_count = self._count(header_row)
        for check in range(header_row - 1, max(0, header_row - 3), -1):
            vals = self._filled(check)
            if len(vals) < 2:
                continue
            if len(vals) < header_count and all(isinstance(v, str) for v in vals.values()):
                if max(vals.keys()) > 5:
                    return check
        return None

    def detect_repeats(self, header_row: int) -> List[int]:
        """Find repeated header tokens matching the index keyword (side-by-side)."""
        positions = []
        for col, val in self._filled(header_row).items():
            if val and str(val).strip().lower() == self.index_keyword:
                positions.append(col)
        return sorted(positions)

    def detect_sections(self, header_row: int, section_row: Optional[int],
                        repeats: List[int]) -> List[Dict[str, str]]:
        sections = []
        note_re = re.compile(self.s.note_pattern) if self.s.note_pattern else None
        if section_row:
            svals = self._filled(section_row)
            starts = sorted(svals.keys())
            for i, sc in enumerate(starts):
                name = str(svals[sc]).strip()
                if note_re:
                    name = note_re.sub("", name).strip()
                end = (starts[i + 1] - 1) if i + 1 < len(starts) else self._max_col_in_row(header_row)
                actual_end = sc
                for c in range(sc, end + 1):
                    if self.reader.cell(self.ws, header_row, c) is not None:
                        actual_end = c
                sections.append({"name": name or f"Section_{i+1}",
                                 "start_col": get_column_letter(sc),
                                 "end_col": get_column_letter(actual_end)})
        elif len(repeats) > 1:
            for i, rc in enumerate(repeats):
                end = (repeats[i + 1] - 1) if i + 1 < len(repeats) else self._max_col_in_row(header_row)
                actual_end = rc
                for c in range(rc, end + 1):
                    if self.reader.cell(self.ws, header_row, c) is not None:
                        actual_end = c
                sections.append({"name": f"Section_{i+1}",
                                 "start_col": get_column_letter(rc),
                                 "end_col": get_column_letter(actual_end)})
        return sections

    def detect_stacked(self) -> List[Dict[str, Any]]:
        tables, in_data = [], False
        for r in range(1, min(self.s.max_scan_rows, self.max_row) + 1):
            filled = self._count(r)
            if filled == 0:
                in_data = False
                continue
            if filled == 1 and not in_data:
                val = next((str(v) for v in self._filled(r).values()), None)
                if val and len(val) > 10 and r + 1 <= self.max_row and self._count(r + 1) >= 2:
                    tables.append({"title_row": r, "header_row": r + 1, "name": val[:50]})
                    in_data = True
                continue
            if filled >= 2:
                in_data = True
        return tables if len(tables) > 1 else []

    def detect_start_col(self, header_row: int) -> str:
        for c in range(1, min(5, self.max_col) + 1):
            if self.reader.cell(self.ws, header_row, c) is not None:
                return get_column_letter(c)
        return "A"

    def detect_empty_columns(self, header_row: int) -> List[str]:
        empties = []
        end = self._max_col_in_row(header_row)
        scan = min(self.max_row, header_row + 200)
        for c in range(1, end + 1):
            if self.reader.cell(self.ws, header_row, c) is None:
                if not any(self.reader.cell(self.ws, r, c) is not None
                           for r in range(header_row + 1, scan + 1)):
                    empties.append(get_column_letter(c))
        return empties

    def get_headers(self, header_row: int, start_col: int = 1,
                    end_col: Optional[int] = None) -> List[Dict[str, Optional[str]]]:
        if end_col is None:
            end_col = self._max_col_in_row(header_row)
        note_re = re.compile(self.s.note_pattern) if self.s.note_pattern else None
        out = []
        for c in range(start_col, end_col + 1):
            val = self.reader.cell(self.ws, header_row, c)
            if val is not None:
                cleaned = re.sub(r"\s+", " ", str(val).replace("\n", " ")).strip()
                if note_re:
                    cleaned = note_re.sub("", cleaned).strip()
                out.append({"col": get_column_letter(c), "name": cleaned})
            else:
                out.append({"col": get_column_letter(c), "name": None})
        return out

    def analyse(self) -> Dict[str, Any]:
        result: Dict[str, Any] = {
            "sheet_name": self.name,
            "dimensions": f"{self.max_row} rows x {self.max_col} cols",
            "last_data_row": self._last_data_row(),
        }
        header_row, candidates = self.detect_header_row()
        if header_row is None:
            result.update(status="NO_HEADER_DETECTED",
                          recommendation="Skip (metadata/cover) or define manually")
            return result
        result.update(header_row=header_row, header_candidates=candidates,
                      data_start_row=header_row + 1,
                      actual_max_col=self._max_col_in_row(header_row))
        start_col = self.detect_start_col(header_row)
        if start_col != "A":
            result["start_col"] = start_col
        sc_idx = _col_to_idx(start_col)
        result["headers"] = self.get_headers(header_row, sc_idx)
        empties = self.detect_empty_columns(header_row)
        if empties:
            result["empty_columns"] = empties
        repeats = self.detect_repeats(header_row)
        result["repeat_columns"] = len(repeats)
        if len(repeats) > 1:
            section_row = self.detect_section_header_row(header_row)
            result["section_header_row"] = section_row
            result["sections"] = self.detect_sections(header_row, section_row, repeats)
            result["table_type"] = "side_by_side"
            result["recommendation"] = f"side_by_side with {len(result['sections'])} sections"
        else:
            stacked = self.detect_stacked()
            if stacked:
                result.update(stacked_tables=stacked, table_type="stacked",
                              recommendation=f"stacked with {len(stacked)} tables")
            else:
                result.update(table_type="simple", recommendation="simple (single table)")
        return result


def build_config(analyses: Dict[str, Dict[str, Any]], config_name: str,
                 source_file: str, index_keyword: str = "Year",
                 section_column_name: str = "Section") -> Dict[str, Any]:
    """Assemble a v2 config from :class:`SheetAnalyser` results."""
    config: Dict[str, Any] = {
        "config_name": config_name,
        "config_version": "2.0",
        "engine_version": ENGINE_VERSION,
        "description": f"Auto-generated config for {os.path.basename(source_file)}",
        "source_info": {
            "original_file": os.path.basename(source_file),
            "generated_by": "Excel Config Generator Wizard",
            "generated_at": utc_now_iso(),
        },
        "settings": {
            "null_markers": list(DEFAULT_NULL_MARKERS),
            "note_pattern": DEFAULT_NOTE_PATTERN,
            "type_inference": "safe",
            "stop_on_blank_rows": 1,
            "resolve_merged_cells": True,
            "dedupe_headers": True,
            "on_error": "continue",
        },
        "sheets": [],
    }
    for name, a in analyses.items():
        if a.get("status") == "NO_HEADER_DETECTED":
            continue
        tt = a.get("table_type", "simple")
        sheet_config: Dict[str, Any] = {
            "sheet_name": name,
            "table_type": tt,
            "description": f"Auto-detected: {a.get('recommendation', '')}",
            "output_name": name.lower().replace(" ", "_").replace(".", "_"),
            "header_row": a["header_row"],
            "data_start_row": a["data_start_row"],
            "index_columns": [],
        }
        if a.get("start_col", "A") != "A":
            sheet_config["start_col"] = a["start_col"]
        if tt == "side_by_side" and a.get("sections"):
            if a.get("section_header_row"):
                sheet_config["section_header_row"] = a["section_header_row"]
            sheet_config["section_column_name"] = section_column_name
            sheet_config["index_columns"] = [index_keyword]
            sheet_config["unpivot"] = False
            sheet_config["unpivot_var_name"] = "Variable"
            sheet_config["unpivot_value_name"] = "Value"
            sheet_config["exclude_from_unpivot"] = ["Total"]
            sheet_config["sections"] = a["sections"]
        if tt == "stacked" and a.get("stacked_tables"):
            tables = []
            for i, t in enumerate(a["stacked_tables"]):
                tbl = {"name": t.get("name", f"Table_{i+1}")[:50],
                       "header_row": t["header_row"],
                       "data_start_row": t["header_row"] + 1}
                if i + 1 < len(a["stacked_tables"]):
                    tbl["data_end_row"] = a["stacked_tables"][i + 1]["title_row"] - 1
                tables.append(tbl)
            sheet_config["tables"] = tables
            sheet_config.pop("header_row", None)
            sheet_config.pop("data_start_row", None)
        # NB: no auto-rename is emitted. The engine already normalises header
        # whitespace/newlines and strips note markers, so long-but-valid names
        # (e.g. "Consumption per unit of output index (2000 = 100)") are kept
        # intact rather than being truncated. Add explicit `rename_columns`
        # entries by hand for business-friendly names.
        config["sheets"].append(sheet_config)
    # De-duplicate output names (suffix on collision).
    seen: Dict[str, int] = {}
    for sc in config["sheets"]:
        base = sc["output_name"]
        if base in seen:
            seen[base] += 1
            sc["output_name"] = f"{base}_{seen[base]}"
        else:
            seen[base] = 1
    return config


# --------------------------------------------------------------------------- #
# Writer
# --------------------------------------------------------------------------- #
class ExtractionWriter:
    """Writes extracted tables and a rich manifest. Uses atomic temp+replace."""

    def __init__(self, config: Dict[str, Any], run_id: str, config_hash: str):
        self.config = config
        self.run_id = run_id
        self.config_hash = config_hash

    @staticmethod
    def _profile(df: pd.DataFrame) -> Dict[str, Any]:
        return {
            "rows": int(len(df)),
            "columns": int(len(df.columns)),
            "column_names": [str(c) for c in df.columns],
            "dtypes": {str(c): str(t) for c, t in df.dtypes.items()},
            "null_counts": {str(c): int(df[c].isna().sum()) for c in df.columns},
        }

    def _atomic_write(self, df: pd.DataFrame, path: str, fmt: str) -> None:
        tmp = f"{path}.{uuid.uuid4().hex}.tmp"
        try:
            if fmt == "parquet":
                df.to_parquet(tmp, index=False, engine="pyarrow")
            else:
                df.to_csv(tmp, index=False)
            os.replace(tmp, path)
        finally:
            if os.path.exists(tmp):
                try:
                    os.remove(tmp)
                except OSError:
                    pass

    def save_all(self, results: Dict[str, "TableResult"], output_path: str,
                 output_format: str, source_file: str) -> Dict[str, Any]:
        output_path = normalize_local_path(output_path)
        os.makedirs(output_path, exist_ok=True)
        manifest_tables = []
        for name, res in results.items():
            if res.status == "error":
                manifest_tables.append({"table": name, "sheet": res.sheet_name,
                                        "status": "error", "error": res.error})
                continue
            if res.dataframe.empty:
                manifest_tables.append({"table": name, "sheet": res.sheet_name,
                                        "status": "empty"})
                continue
            file_name = f"{name}.{output_format}"
            self._atomic_write(res.dataframe, os.path.join(output_path, file_name),
                               output_format)
            manifest_tables.append({"table": name, "sheet": res.sheet_name,
                                    "file": file_name, "status": "ok",
                                    "duration_ms": round(res.duration_ms, 1),
                                    **self._profile(res.dataframe)})
            logger.info("Saved %s (%d rows)", file_name, res.rows)

        manifest = {
            "run_id": self.run_id,
            "config_name": self.config["config_name"],
            "config_version": self.config.get("config_version", "unknown"),
            "engine_version": ENGINE_VERSION,
            "config_hash": self.config_hash,
            "extracted_at": utc_now_iso(),
            "source_file": source_file,
            "output_format": output_format,
            "tables": manifest_tables,
            "total_tables_ok": sum(1 for t in manifest_tables if t.get("status") == "ok"),
            "total_tables_error": sum(1 for t in manifest_tables if t.get("status") == "error"),
            "total_rows": sum(t.get("rows", 0) for t in manifest_tables),
        }
        with open(os.path.join(output_path, "_extraction_manifest.json"), "w") as f:
            json.dump(manifest, f, indent=2)
        logger.info("Manifest written to %s/_extraction_manifest.json", output_path)
        return manifest
